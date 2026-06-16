#!/usr/bin/env python3
"""
Validerer sider fra et Excel-ark og lager en rapport (JSON + interaktiv HTML).

For hver URL i kolonnene "url" og "ny-url":
  - HTTP-status (lastet siden OK?)
  - a11y-sjekk med axe-core (WCAG-brudd, gruppert etter alvorlighetsgrad)
  - SEO/meta: title, meta description, lang, antall h1, manglende alt, canonical
  - Brutte lenker: alle <a href> på siden sjekkes for 4xx/5xx/feil

Resultat:
  report.json  – alle data
  report.html  – selvstendig side med søk + filtrering (åpnes ved dobbeltklikk)

Bruker samme venv som screenshot-scriptet. Ingen ny installasjon nødvendig
(axe-core lastes ned automatisk første gang og caches lokalt).

Bruk:
    python3 validate_pages.py status_fra_sis.xlsx
    python3 validate_pages.py status_fra_sis.xlsx --only ny
    python3 validate_pages.py status_fra_sis.xlsx --internal-only
    python3 validate_pages.py status_fra_sis.xlsx --skip-links --concurrency 6
"""

import argparse
import asyncio
import datetime
import json
import re
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path
from urllib.parse import urldefrag, urljoin, urlparse

from openpyxl import load_workbook
from playwright.async_api import async_playwright

AXE_VERSION = "4.10.2"
AXE_URL = f"https://cdnjs.cloudflare.com/ajax/libs/axe-core/{AXE_VERSION}/axe.min.js"

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)


def read_pages(excel_path: Path, only: str | None):
    """Returnerer liste av (radnr, kolonne, url). Kolonne = 'gammel' eller 'ny'."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_raw = [str(c).strip() if c is not None else "" for c in rows[0]]
    header = [h.lower() for h in header_raw]
    try:
        i_url = header.index("url")
        i_ny = header.index("ny-url")
    except ValueError:
        raise SystemExit(
            f'Fant ikke kolonnene "url" og "ny-url". Overskrifter: {header}'
        )

    extra_cols = [
        (idx, header_raw[idx]) for idx in range(len(header_raw))
        if idx not in (i_url, i_ny) and header_raw[idx]
    ]

    pages = []
    for r_i, r in enumerate(rows[1:], start=1):
        url = r[i_url] if i_url < len(r) else None
        ny = r[i_ny] if i_ny < len(r) else None
        extra = {
            name: (str(r[idx]).strip() if idx < len(r) and r[idx] is not None else "")
            for idx, name in extra_cols
        }
        if url and only in (None, "url", "gammel"):
            pages.append((r_i, "gammel", str(url).strip(), extra))
        if ny and only in (None, "ny"):
            pages.append((r_i, "ny", str(ny).strip(), extra))
    return pages


def _localname(tag: str) -> str:
    """'{ns}loc' -> 'loc' (sitemap-XML har namespace)."""
    return tag.rsplit("}", 1)[-1].lower()


def _fetch_sitemap_locs(url: str, depth: int = 0, seen: set | None = None) -> list[str]:
    """Returnerer alle <loc>-URL-er. Følger sitemap-index rekursivt (maks 3 nivå)."""
    seen = seen if seen is not None else set()
    if url in seen or depth > 3:
        return []
    seen.add(url)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read()
    except Exception as e:
        print(f"  ADVARSEL: klarte ikke hente sitemap {url}: {e}")
        return []
    try:
        root = ET.fromstring(raw)
    except ET.ParseError as e:
        print(f"  ADVARSEL: kunne ikke parse sitemap {url}: {e}")
        return []

    is_index = _localname(root.tag) == "sitemapindex"
    locs = [
        el.text.strip()
        for el in root.iter()
        if _localname(el.tag) == "loc" and el.text and el.text.strip()
    ]
    if is_index:
        out: list[str] = []
        for child in locs:
            out.extend(_fetch_sitemap_locs(child, depth + 1, seen))
        return out
    return locs


def read_sitemap(sitemap_url: str, limit: int | None):
    """Henter URL-er fra en sitemap.xml. Returnerer samme form som read_pages:
    (radnr, kolonne, url, extra). Hver URL er sin egen side (ingen gammel/ny-par)."""
    locs = _fetch_sitemap_locs(sitemap_url)
    # dedupe, behold rekkefølge
    seen: set[str] = set()
    urls = [u for u in locs if not (u in seen or seen.add(u))]
    if limit and limit > 0:
        urls = urls[:limit]
    return [(i, "ny", u, {}) for i, u in enumerate(urls, start=1)]


async def crawl_site(base_url: str, limit: int | None):
    """Fallback når sitemap mangler: bredde-først-traversering fra base_url med
    full JS-rendring (Playwright), så vi også fanger lenker som legges inn av
    JavaScript (SPA-er). Følger interne lenker (samme origin). Samme form som
    read_sitemap. Uten --limit settes et hardt tak på 100 sider."""
    max_pages = limit if (limit and limit > 0) else 100

    # Behandle www.host og host som samme nettsted (svært vanlig at kanonisk vert
    # er den ene, men lenkene peker til den andre).
    def hostkey(netloc: str) -> str:
        return netloc.lower().removeprefix("www.")

    origin = hostkey(urlparse(base_url).netloc)
    seen: set[str] = set()
    order: list[str] = []
    queue: list[str] = [base_url]

    def norm(u: str) -> str:
        u = urldefrag(u)[0]
        p = urlparse(u)
        path = p.path or "/"
        if path != "/" and path.endswith("/"):
            path = path[:-1]
        q = f"?{p.query}" if p.query else ""
        return f"{p.scheme}://{hostkey(p.netloc)}{path}{q}"

    async with async_playwright() as pw:
        browser = await pw.chromium.launch()
        page = await browser.new_page(user_agent=UA)
        while queue and len(order) < max_pages:
            url = queue.pop(0)
            key = norm(url)
            if key in seen:
                continue
            seen.add(key)
            try:
                await page.goto(url, wait_until="load", timeout=30000)
                await page.wait_for_timeout(800)  # la hydrering legge til lenker
                hrefs = await page.eval_on_selector_all(
                    "a[href]", "els => els.map(e => e.href)"
                )
            except Exception:
                continue
            order.append(url)
            for h in hrefs:
                link = urldefrag(h)[0]
                p = urlparse(link)
                if (
                    p.scheme in ("http", "https")
                    and hostkey(p.netloc) == origin
                    and norm(link) not in seen
                ):
                    queue.append(link)
        await browser.close()
    print(f"Crawl fant {len(order)} sider (origin: {origin}).")
    return [(i, "ny", u, {}) for i, u in enumerate(order, start=1)]


def get_axe_source(cache: Path) -> str | None:
    """Henter axe-core-kildekoden (caches lokalt etter første nedlasting)."""
    if cache.exists():
        return cache.read_text(encoding="utf-8")
    try:
        print(f"Laster ned axe-core {AXE_VERSION} (engangsjobb) ...")
        req = urllib.request.Request(AXE_URL, headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=30) as resp:
            src = resp.read().decode("utf-8")
        cache.write_text(src, encoding="utf-8")
        return src
    except Exception as e:
        print(f"  ADVARSEL: klarte ikke hente axe-core ({e}). a11y hoppes over.")
        return None


JS_META = """() => {
    const metaName = n => { const e = document.querySelector(`meta[name="${n}"]`); return e ? e.getAttribute('content') : null; };
    const metaProp = p => { const e = document.querySelector(`meta[property="${p}"]`); return e ? e.getAttribute('content') : null; };

    const title = document.title || null;
    const desc = metaName('description');
    const lang = document.documentElement.getAttribute('lang') || null;
    const h1s = [...document.querySelectorAll('h1')]
        .map(h => h.textContent.trim()).filter(Boolean);
    const canEl = document.querySelector('link[rel="canonical"]');
    const canonical = canEl ? canEl.href : null;
    const imgs = [...document.querySelectorAll('img')];
    const noAltAttr = imgs.filter(i => !i.hasAttribute('alt')).length;

    const og = {
        title: metaProp('og:title'), description: metaProp('og:description'),
        image: metaProp('og:image'), type: metaProp('og:type'),
        url: metaProp('og:url'), site_name: metaProp('og:site_name')
    };
    const twitter_card = metaName('twitter:card');
    const robots_meta = metaName('robots');
    const viewport = metaName('viewport');

    const hreflang = [...document.querySelectorAll('link[rel="alternate"][hreflang]')]
        .map(l => ({ lang: l.getAttribute('hreflang'), href: l.href }));

    const jsonld = [];
    for (const s of document.querySelectorAll('script[type="application/ld+json"]')) {
        try {
            const data = JSON.parse(s.textContent);
            const arr = Array.isArray(data) ? data : [data];
            for (const o of arr) {
                if (o && o['@type']) jsonld.push(Array.isArray(o['@type']) ? o['@type'].join(', ') : o['@type']);
            }
        } catch (e) { jsonld.push('(ugyldig JSON-LD)'); }
    }

    const levels = [...document.querySelectorAll('h1,h2,h3,h4,h5,h6')].map(h => +h.tagName[1]);
    let heading_skips = 0, prev = 0;
    for (const lvl of levels) { if (prev && lvl > prev + 1) heading_skips++; prev = lvl; }

    const renderedText = (document.body.innerText || '').replace(/\\s+/g, ' ').trim();

    return {
        title,
        title_length: title ? title.length : 0,
        meta_description: desc,
        description_length: desc ? desc.length : 0,
        lang,
        h1_count: h1s.length,
        h1_texts: h1s.slice(0, 5),
        canonical,
        images_total: imgs.length,
        images_missing_alt: noAltAttr,
        og,
        twitter_card,
        robots_meta,
        viewport,
        hreflang,
        jsonld,
        heading_count: levels.length,
        heading_skips,
        rendered_text_len: renderedText.length,
        word_count: renderedText ? renderedText.split(' ').length : 0
    };
}"""

# Ytelse/vekt via Performance-API-et + bilde-dimensjoner. Vekt er tilnærmet:
# transferSize er 0 for kryss-origin-ressurser uten Timing-Allow-Origin.
JS_PERF = """() => {
    const nav = performance.getEntriesByType('navigation')[0] || {};
    const res = performance.getEntriesByType('resource') || [];
    const byType = {};
    let total = 0;
    for (const r of res) {
        const sz = r.transferSize || r.encodedBodySize || 0;
        total += sz;
        const t = r.initiatorType || 'other';
        byType[t] = (byType[t] || 0) + sz;
    }
    const heaviest = res
        .map(r => ({ url: r.name, bytes: r.transferSize || r.encodedBodySize || 0, type: r.initiatorType || 'other' }))
        .filter(r => r.bytes > 0)
        .sort((a, b) => b.bytes - a.bytes)
        .slice(0, 6);
    const dpr = window.devicePixelRatio || 1;
    const oversized = [];
    for (const im of document.querySelectorAll('img')) {
        const dispW = Math.round(im.clientWidth * dpr);
        if (im.naturalWidth && im.clientWidth > 1 && im.naturalWidth > dispW * 1.5) {
            oversized.push({ src: im.currentSrc || im.src, naturalW: im.naturalWidth, displayW: im.clientWidth });
        }
    }
    return {
        ttfb_ms: Math.round(nav.responseStart || 0),
        dcl_ms: Math.round(nav.domContentLoadedEventEnd || 0),
        load_ms: Math.round(nav.loadEventEnd || 0),
        weight_total: total,
        weight_img: byType.img || 0,
        weight_js: byType.script || 0,
        weight_css: (byType.link || 0) + (byType.css || 0),
        resource_count: res.length,
        dom_nodes: document.getElementsByTagName('*').length,
        img_oversized: oversized.length,
        img_oversized_examples: oversized.slice(0, 5),
        heaviest,
    };
}"""

JS_LINKS = """() => {
    const out = new Map();
    for (const a of document.querySelectorAll('a[href]')) {
        let href;
        try { href = new URL(a.getAttribute('href'), location.href).href; }
        catch { continue; }
        if (!href.startsWith('http')) continue;
        if (!out.has(href)) out.set(href, (a.textContent || '').trim().slice(0, 80));
    }
    return [...out.entries()].map(([url, text]) => ({ url, text }));
}"""


def summarize_axe(results: dict) -> dict:
    violations = results.get("violations", []) if results else []
    incomplete = results.get("incomplete", []) if results else []
    by_impact = {"critical": 0, "serious": 0, "moderate": 0, "minor": 0}
    v_list = []
    for v in violations:
        imp = v.get("impact") or "minor"
        by_impact[imp] = by_impact.get(imp, 0) + 1
        nodes = v.get("nodes", [])
        v_list.append(
            {
                "id": v.get("id"),
                "impact": imp,
                "help": v.get("help"),
                "helpUrl": v.get("helpUrl"),
                "nodes": len(nodes),
                "targets": [
                    (n.get("target") or [""])[0] for n in nodes[:5]
                ],
            }
        )
    # sorter etter alvorlighetsgrad
    order = {"critical": 0, "serious": 1, "moderate": 2, "minor": 3}
    v_list.sort(key=lambda x: order.get(x["impact"], 9))
    # "incomplete" = axe kunne ikke avgjøre automatisk -> krever manuell sjekk
    inc_list = [
        {"id": i.get("id"), "help": i.get("help"), "helpUrl": i.get("helpUrl"),
         "nodes": len(i.get("nodes", []))}
        for i in incomplete
    ]
    return {
        "violation_count": len(violations),
        "by_impact": by_impact,
        "violations": v_list,
        "incomplete_count": len(incomplete),
        "incomplete": inc_list[:15],
    }


def compute_seo(meta: dict) -> list:
    """Returnerer liste av {level, key, msg}. level: 'fail' | 'warn' | 'ok'."""
    out = []
    if not meta:
        return out
    t, tl = meta.get("title"), meta.get("title_length", 0)
    if not t:
        out.append({"level": "fail", "key": "title-missing", "msg": "Mangler <title>"})
    elif tl > 60:
        out.append({"level": "warn", "key": "title-long", "msg": f"Title er lang ({tl} tegn) – klippes ofte i søk (~60)"})
    elif tl < 15:
        out.append({"level": "warn", "key": "title-short", "msg": f"Title er kort ({tl} tegn)"})
    else:
        out.append({"level": "ok", "key": "title-ok", "msg": f"Title OK ({tl} tegn)"})

    d, dl = meta.get("meta_description"), meta.get("description_length", 0)
    if not d:
        out.append({"level": "warn", "key": "desc-missing", "msg": "Mangler meta description"})
    elif dl > 300:
        out.append({"level": "fail", "key": "desc-long",
                    "msg": f"Meta description altfor lang ({dl} tegn) – trolig autogenerert fra brødtekst. Sett en egen på ~120–155 tegn."})
    elif dl > 160:
        out.append({"level": "warn", "key": "desc-long",
                    "msg": f"Meta description for lang ({dl} tegn) – klippes i søk (anbefalt ~50–160)"})
    elif dl < 50:
        out.append({"level": "warn", "key": "desc-short", "msg": f"Meta description kort ({dl} tegn, anbefalt ~50–160)"})
    else:
        out.append({"level": "ok", "key": "desc-ok", "msg": f"Meta description OK ({dl} tegn)"})

    if not meta.get("lang"):
        out.append({"level": "fail", "key": "lang-missing", "msg": "Mangler lang-attributt på <html>"})
    if not meta.get("canonical"):
        out.append({"level": "warn", "key": "canonical-missing", "msg": "Mangler canonical-lenke"})
    if not meta.get("viewport"):
        out.append({"level": "warn", "key": "viewport-missing", "msg": "Mangler viewport-meta (mobil)"})

    h1 = meta.get("h1_count", 0)
    if h1 == 0:
        out.append({"level": "fail", "key": "h1-none", "msg": "Ingen h1 på siden"})
    elif h1 > 1:
        out.append({"level": "warn", "key": "h1-many", "msg": f"{h1} h1-overskrifter (bør være én)"})

    if meta.get("heading_skips", 0):
        out.append({"level": "warn", "key": "heading-skips",
                    "msg": f"Hopp i overskriftsnivåer ({meta['heading_skips']} steder) – f.eks. h2→h4"})

    robots = (meta.get("robots_meta") or "").lower()
    if "noindex" in robots:
        out.append({"level": "fail", "key": "noindex", "msg": "noindex satt – siden indekseres IKKE av søkemotorer"})
    if "nofollow" in robots:
        out.append({"level": "warn", "key": "nofollow", "msg": "nofollow satt i robots-meta"})

    og = meta.get("og") or {}
    if not og.get("image"):
        out.append({"level": "warn", "key": "og-image",
                    "msg": "Mangler og:image – ingen forhåndsvisningsbilde ved deling/AI-kort"})
    miss_meta = [k for k in ("title", "description") if not og.get(k)]
    if miss_meta:
        out.append({"level": "warn", "key": "og-meta",
                    "msg": "Mangler Open Graph: " + ", ".join("og:" + m for m in miss_meta)})
    if not meta.get("twitter_card"):
        out.append({"level": "warn", "key": "twitter-missing", "msg": "Mangler twitter:card"})

    return out


def compute_geo(meta: dict, ssr: dict, markdown: dict, site_robots: dict) -> dict:
    """GEO = Generative Engine Optimization: synlighet for AI-søk/chatbotter."""
    tips = []
    signals = {}
    md = meta or {}

    jsonld = md.get("jsonld") or []
    signals["jsonld"] = jsonld
    if not jsonld:
        tips.append("Legg til JSON-LD strukturert data (Organization, WebPage, FAQPage …). AI-motorer bruker dette aktivt for å forstå og sitere innhold.")

    # Server-rendret innhold? AI-crawlere kjører som regel ikke JS.
    rendered = md.get("rendered_text_len", 0)
    ssr_len = (ssr or {}).get("text_len", 0)
    signals["rendered_text_len"] = rendered
    signals["ssr_text_len"] = ssr_len
    js_dependent = bool(rendered and ssr_len < rendered * 0.5)
    signals["js_dependent"] = js_dependent
    if js_dependent:
        tips.append(f"Mye innhold lastes via JS (server: ~{ssr_len} tegn vs. rendret: ~{rendered}). Mange AI-crawlere (og enklere botter) ser bare server-HTML – vurder SSR/SSG for hovedinnholdet.")

    wc = md.get("word_count", 0)
    signals["word_count"] = wc
    if wc and wc < 100:
        tips.append(f"Tynt tekstinnhold (~{wc} ord). AI-motorer favoriserer substansielt, faktatett innhold med tydelig struktur.")

    signals["markdown_available"] = (markdown or {}).get("available", False)
    if markdown and not markdown.get("available"):
        tips.append("Markdown-versjon ble ikke detektert (verken via Accept: text/markdown eller .md). Ren markdown er gull for LLM-er – verdt å verifisere at featuren svarer riktig.")

    # robots.txt blokkerer AI-botter?
    if site_robots and site_robots.get("ai_bots"):
        blocked = [b for b, ok in site_robots["ai_bots"].items() if ok is False]
        signals["ai_bots_blocked"] = blocked
        if blocked:
            tips.append("robots.txt blokkerer AI-botter: " + ", ".join(blocked) + ". Hvis du ønsker synlighet i AI-søk, vurder å åpne for dem.")

    if not md.get("meta_description"):
        tips.append("Meta description mangler – brukes ofte som sammendrag av både søkemotorer og AI.")

    return {"signals": signals, "tips": tips}


def html_to_text_len(html: str) -> int:
    """Grov tekstlengde fra rå HTML (uten å kjøre JS) – for SSR-sjekk."""
    h = re.sub(r"<script[\s\S]*?</script>", " ", html, flags=re.I)
    h = re.sub(r"<style[\s\S]*?</style>", " ", h, flags=re.I)
    h = re.sub(r"<[^>]+>", " ", h)
    h = re.sub(r"\s+", " ", h).strip()
    return len(h)


async def check_ssr(context, url):
    """Henter rå server-HTML (ingen JS) og måler tekstmengde."""
    try:
        r = await context.request.get(
            url, timeout=20000, headers={"User-Agent": UA}
        )
        html = await r.text()
        return {"status": r.status, "text_len": html_to_text_len(html)}
    except Exception as e:
        return {"error": str(e), "text_len": 0}


async def check_markdown(context, url):
    """Sjekker om siden kan serveres som markdown (Accept-header + .md)."""
    out = {"accept_header": None, "dot_md": None, "available": False}
    try:
        r = await context.request.get(
            url, timeout=15000,
            headers={"User-Agent": UA, "Accept": "text/markdown"},
        )
        ct = (r.headers.get("content-type") or "").lower()
        out["accept_header"] = {"status": r.status, "content_type": ct}
        if "markdown" in ct:
            out["available"] = True
    except Exception as e:
        out["accept_header"] = {"error": str(e)}
    try:
        md_url = url.split("?")[0].rstrip("/") + ".md"
        r = await context.request.get(
            md_url, timeout=15000, headers={"User-Agent": UA}
        )
        ct = (r.headers.get("content-type") or "").lower()
        out["dot_md"] = {"status": r.status, "content_type": ct, "url": md_url}
        if r.status == 200 and "markdown" in ct:
            out["available"] = True
    except Exception as e:
        out["dot_md"] = {"error": str(e)}
    return out


JS_FOCUS_PROBE = """() => {
    const el = document.activeElement;
    if (!el || el === document.body || el === document.documentElement) return null;
    window.__kb = window.__kb || [];
    const seen = window.__kb.includes(el);
    const sameAsLast = window.__kbLast === el;
    window.__kbLast = el;
    if (!seen) window.__kb.push(el);
    const cs = getComputedStyle(el);
    const r = el.getBoundingClientRect();
    const inVP = r.width > 0 && r.height > 0 && r.bottom > 0 && r.right > 0 &&
                 r.top < innerHeight && r.left < innerWidth;
    let sel = el.tagName.toLowerCase();
    if (el.id) sel += '#' + el.id;
    else if (typeof el.className === 'string' && el.className.trim())
        sel += '.' + el.className.trim().split(/\\s+/).slice(0, 2).join('.');
    const name = (el.getAttribute('aria-label') || el.textContent || el.value ||
                  el.getAttribute('alt') || '').trim().replace(/\\s+/g, ' ').slice(0, 60);
    const ow = parseFloat(cs.outlineWidth) || 0;
    const hasOutline = cs.outlineStyle !== 'none' && ow > 0;
    const hasShadow = !!cs.boxShadow && cs.boxShadow !== 'none';
    const ariaHidden = !!el.closest('[aria-hidden="true"]');
    return {
        tag: el.tagName.toLowerCase(), sel, name,
        tabindex: el.getAttribute('tabindex'),
        rect: {w: Math.round(r.width), h: Math.round(r.height)},
        inViewport: inVP, hasOutline, hasShadow, ariaHidden, seen, sameAsLast
    };
}"""

# Interaktive elementer som IKKE er tastaturtilgjengelige
JS_UNREACHABLE = """() => {
    const focusableTags = new Set(['A','BUTTON','INPUT','SELECT','TEXTAREA','SUMMARY']);
    const roles = new Set(['button','link','tab','menuitem','menuitemcheckbox','menuitemradio',
                           'checkbox','radio','switch','option','slider']);
    const out = [];
    for (const el of document.querySelectorAll('[role],[onclick]')) {
        const role = (el.getAttribute('role') || '').toLowerCase();
        if (!(roles.has(role) || el.hasAttribute('onclick'))) continue;
        if (focusableTags.has(el.tagName)) continue;
        const ti = el.getAttribute('tabindex');
        if (ti !== null && parseInt(ti) >= 0) continue;
        const r = el.getBoundingClientRect();
        if (r.width === 0 && r.height === 0) continue;       // skjult
        if (el.closest('[inert]')) continue;
        let sel = el.tagName.toLowerCase();
        if (el.id) sel += '#' + el.id;
        else if (typeof el.className === 'string' && el.className.trim())
            sel += '.' + el.className.trim().split(/\\s+/).slice(0, 2).join('.');
        out.push({ sel, role: role || '(onclick)',
                   name: (el.textContent || '').trim().replace(/\\s+/g, ' ').slice(0, 50) });
        if (out.length >= 30) break;
    }
    return out;
}"""


async def capture_shot(page, path):
    """Tar et full-height JPEG-skjermbilde (skjuler cookie-banner, scroller for lazy-last)."""
    try:
        await page.evaluate(
            """() => {
                for (const s of ['#CybotCookiebotDialog','#CybotCookiebotDialogBodyUnderlay','#CookiebotWidget'])
                    document.querySelectorAll(s).forEach(e => e.remove());
                document.documentElement.style.overflow = ''; document.body.style.overflow = '';
            }"""
        )
        await page.evaluate(
            """async () => { await new Promise(r => {
                let t = 0; const s = Math.max(800, innerHeight);
                const i = setInterval(() => { scrollBy(0, s); t += s;
                    if (t >= document.body.scrollHeight) { clearInterval(i); scrollTo(0, 0); r(); } }, 30);
            }); }"""
        )
        await page.wait_for_timeout(250)
        await page.screenshot(path=str(path), full_page=True, type="jpeg", quality=72)
        return True
    except Exception:
        return False


def stitch_pair(left_path, right_path, out_path):
    """Syr gammel + ny sammen til ett bilde side om side, toppjustert, med etiketter."""
    from PIL import Image, ImageDraw, ImageFont

    left = Image.open(left_path).convert("RGB")
    right = Image.open(right_path).convert("RGB")
    gap, label_h = 2, 44
    width = left.width + gap + right.width
    height = max(left.height, right.height) + label_h
    canvas = Image.new("RGB", (width, height), "white")
    canvas.paste(left, (0, label_h))
    canvas.paste(right, (left.width + gap, label_h))
    draw = ImageDraw.Draw(canvas)
    draw.rectangle([left.width, label_h, left.width + gap, height], fill=(200, 200, 200))
    try:
        font = ImageFont.truetype("DejaVuSans-Bold.ttf", 24)
    except Exception:
        font = ImageFont.load_default()
    draw.text((10, 12), "Gammel (url)", fill=(20, 20, 20), font=font)
    draw.text((left.width + gap + 10, 12), "Ny (ny-url)", fill=(20, 20, 20), font=font)
    canvas.save(out_path, "JPEG", quality=78)


async def check_keyboard(page, max_tabs):
    """Tabber gjennom siden og vurderer synlig fokus (WCAG 2.4.7) + tab-rekkefølge."""
    try:
        await page.evaluate(
            "() => { try { document.activeElement && document.activeElement.blur(); }"
            " catch(e){} window.__kb=[]; window.__kbLast=null; window.scrollTo(0,0); }"
        )
    except Exception:
        pass

    stops = []
    stuck = 0
    trap = False
    for i in range(max_tabs):
        try:
            await page.keyboard.press("Tab")
            info = await page.evaluate(JS_FOCUS_PROBE)
        except Exception:
            break
        if not info:
            break  # fokus forlot dokumentet (normal slutt)
        if info["sameAsLast"]:
            stuck += 1
            if stuck >= 4:
                trap = True
                break
            continue
        stuck = 0
        if info["seen"]:
            break  # vi har syklet tilbake til et allerede besøkt element
        stops.append(info)

    # interaktive elementer som ikke kan nås med tastatur (egen DOM-skanning)
    try:
        unreachable = await page.evaluate(JS_UNREACHABLE)
    except Exception:
        unreachable = []

    no_focus = [s for s in stops if not s["hasOutline"] and not s["hasShadow"]]
    positive_ti = [
        s for s in stops
        if s["tabindex"] and s["tabindex"].lstrip("-").isdigit() and int(s["tabindex"]) > 0
    ]
    offscreen = [
        s for s in stops
        if not s["inViewport"] and (s["rect"]["w"] > 0 or s["rect"]["h"] > 0)
    ]
    aria_hidden = [s for s in stops if s.get("ariaHidden")]
    skip = None
    if stops:
        nm = (stops[0]["name"] or "").lower()
        skip = {
            "present": any(w in nm for w in ("hopp", "skip", "innhold", "main")),
            "text": stops[0]["name"],
        }
    return {
        "tab_stops": len(stops),
        "trap": trap,
        "skip_link": skip,
        "no_focus_count": len(no_focus),
        "no_focus": [{"tag": s["tag"], "sel": s["sel"], "name": s["name"]} for s in no_focus[:20]],
        "positive_tabindex": [{"sel": s["sel"], "tabindex": s["tabindex"], "name": s["name"]} for s in positive_ti[:20]],
        "offscreen_count": len(offscreen),
        "offscreen": [{"sel": s["sel"], "name": s["name"]} for s in offscreen[:20]],
        "aria_hidden_count": len(aria_hidden),
        "aria_hidden": [{"sel": s["sel"], "name": s["name"]} for s in aria_hidden[:20]],
        "unreachable_count": len(unreachable),
        "unreachable": unreachable[:20],
    }


AI_BOTS = [
    "GPTBot", "OAI-SearchBot", "ChatGPT-User", "ClaudeBot", "anthropic-ai",
    "Claude-Web", "PerplexityBot", "Google-Extended", "CCBot", "Bytespider",
    "Applebot-Extended", "Amazonbot", "Meta-ExternalAgent",
]


async def check_site(context, origin, scheme):
    """Nettsted-nivå: robots.txt (AI-botter + sitemaps) og llms.txt."""
    from urllib.robotparser import RobotFileParser

    base = f"{scheme}://{origin}"
    site = {"origin": origin, "base": base, "robots": None,
            "llms_txt": None, "llms_full_txt": None}

    try:
        r = await context.request.get(base + "/robots.txt", timeout=15000,
                                      headers={"User-Agent": UA})
        ct = (r.headers.get("content-type") or "").lower()
        # SPA-er svarer ofte 200 + text/html på alt (også /robots.txt). En ekte
        # robots.txt er ren tekst – krev at svaret ikke er HTML.
        if r.status == 200 and "html" not in ct:
            txt = await r.text()
            rp = RobotFileParser()
            rp.parse(txt.splitlines())
            ai = {}
            for b in AI_BOTS:
                try:
                    ai[b] = rp.can_fetch(b, base + "/")
                except Exception:
                    ai[b] = None
            site["robots"] = {
                "exists": True,
                "ai_bots": ai,
                "wildcard_allowed": rp.can_fetch("*", base + "/"),
                "sitemaps": rp.site_maps() or [],
            }
        else:
            site["robots"] = {"exists": False, "status": r.status, "content_type": ct}
    except Exception as e:
        site["robots"] = {"error": str(e)}

    for key, path in (("llms_txt", "/llms.txt"), ("llms_full_txt", "/llms-full.txt")):
        try:
            r = await context.request.get(base + path, timeout=15000,
                                          headers={"User-Agent": UA})
            ct = (r.headers.get("content-type") or "").lower()
            # Samme felle: en ekte llms.txt er tekst/markdown, ikke text/html.
            exists = r.status == 200 and "html" not in ct
            site[key] = {"exists": exists, "status": r.status,
                         "content_type": ct, "url": base + path}
        except Exception as e:
            site[key] = {"error": str(e)}

    # Soft-404: svarer nettstedet 200 på en URL som garantert ikke finnes?
    # (SPA-er med catch-all gjør ofte det – dårlig for SEO og skjuler ekte feil.)
    try:
        probe = base + "/qa-monitor-404-probe-zx91kqd7"
        r = await context.request.get(probe, timeout=15000, headers={"User-Agent": UA})
        site["soft_404"] = {"status": r.status, "is_soft_404": r.status == 200, "probe": probe}
    except Exception as e:
        site["soft_404"] = {"error": str(e)}

    return site


def classify_link(status):
    """'ok' = funker, 'broken' = reelt brutt, 'uncertain' = trolig bot-/auth-blokkering."""
    if status is None:
        return "broken"  # ingen respons etter flere forsøk
    if status in (400, 401, 403, 405, 406, 429, 999):
        return "uncertain"  # store sider svarer ofte slik på ikke-nettleser-trafikk
    if status in (404, 410) or status >= 500:
        return "broken"
    return "ok"


def link_ignored(url, patterns):
    """True hvis url matcher et ignore-mønster (eksakt, eller prefiks med avsluttende *)."""
    nu = url.rstrip("/")
    for pat in patterns:
        if pat.endswith("*"):
            if nu.startswith(pat[:-1].rstrip("/")):
                return True
        elif nu == pat.rstrip("/"):
            return True
    return False


async def check_link(context, url, cache, sem):
    """Sjekker én lenke med GET + retry. Returnerer status-kode (int) eller None."""
    if url in cache:
        return cache[url]
    async with sem:
        status = None
        # GET (ikke HEAD) – mange servere avviser/dropper HEAD. Retry for
        # treghet/cold-start på f.eks. dev-miljøer.
        for attempt in range(3):
            try:
                resp = await context.request.get(
                    url,
                    timeout=20000,
                    max_redirects=5,
                    headers={
                        "User-Agent": UA,
                        "Accept": "text/html,application/xhtml+xml,*/*",
                    },
                )
                status = resp.status
                break
            except Exception:
                status = None
                if attempt < 2:
                    await asyncio.sleep(0.6)
        cache[url] = status
        return status


async def analyze(
    context, axe_src, row, column, url, args, link_cache, link_sem, site_robots, extra
):
    page = await context.new_page()
    entry = {
        "row": row,
        "column": column,
        "url": url,
        "status": None,
        "ok": False,
        "load_error": None,
        "meta": None,
        "a11y": None,
        "links": None,
        "seo": None,
        "geo": None,
        "markdown": None,
        "ssr": None,
        "keyboard": None,
        "shot": None,
        "combined_shot": None,
        "extra": extra or {},
    }
    try:
        resp = await page.goto(url, wait_until="load", timeout=args.timeout)
        entry["status"] = resp.status if resp else None
        entry["ok"] = bool(resp and resp.ok)

        entry["meta"] = await page.evaluate(JS_META)
        try:
            perf = await page.evaluate(JS_PERF)
            if isinstance(entry["meta"], dict):
                entry["meta"]["perf"] = perf
        except Exception:
            pass

        if axe_src:
            try:
                await page.add_script_tag(content=axe_src)
                results = await page.evaluate(
                    "async () => await axe.run(document, "
                    "{ resultTypes: ['violations', 'incomplete'] })"
                )
                entry["a11y"] = summarize_axe(results)
            except Exception as e:
                entry["a11y"] = {"error": str(e)}

        if not args.skip_links:
            links = await page.evaluate(JS_LINKS)
            origin = urlparse(url).netloc
            candidates = []
            for ln in links:
                if args.internal_only and urlparse(ln["url"]).netloc != origin:
                    continue
                candidates.append(ln)
            statuses = await asyncio.gather(
                *(check_link(context, ln["url"], link_cache, link_sem)
                  for ln in candidates)
            )
            broken, uncertain, ignored = [], [], []
            for ln, st in zip(candidates, statuses):
                rec = {"url": ln["url"], "text": ln["text"], "status": st}
                if link_ignored(ln["url"], args.ignore_patterns):
                    ignored.append(rec)
                    continue
                kind = classify_link(st)
                if kind == "broken":
                    broken.append(rec)
                elif kind == "uncertain":
                    uncertain.append(rec)
            entry["links"] = {
                "total": len(candidates),
                "broken": broken,
                "uncertain": uncertain,
                "ignored": ignored,
            }

        # SEO-vurdering
        entry["seo"] = compute_seo(entry["meta"])

        # Tastatur / fokus (med mindre slått av)
        if not args.fast and not args.skip_keyboard:
            entry["keyboard"] = await check_keyboard(page, args.max_tabs)

        # Skjermbilde (full høyde) hvis aktivert
        if args.screenshots:
            shot = args.shots_dir / f"{row}_{column}.jpg"
            if await capture_shot(page, shot):
                entry["shot"] = f"shots/{shot.name}"

        # GEO / AI-synlighet (med mindre --fast)
        if not args.fast:
            entry["ssr"] = await check_ssr(context, url)
            entry["markdown"] = await check_markdown(context, url)
        entry["geo"] = compute_geo(
            entry["meta"], entry["ssr"], entry["markdown"], site_robots
        )
    except Exception as e:
        entry["load_error"] = str(e)
    finally:
        await page.close()
    return entry


HTML_TEMPLATE = r"""<!doctype html>
<html lang="nb">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Validerings­rapport</title>
<style>
  :root {
    --bg:#f5f6f8; --card:#fff; --line:#e6e8ec; --line-strong:#d4d8de;
    --text:#1b2230; --muted:#697587; --faint:#9aa3b2;
    --crit:#c01f1f; --serious:#c2410c; --moderate:#9a6b00; --minor:#5b6472;
    --ok:#1a8a4d; --link:#2056d6;
    --radius:12px; --radius-sm:8px;
    --shadow:0 1px 2px rgba(16,24,40,.04), 0 1px 3px rgba(16,24,40,.06);
    --shadow-hover:0 2px 6px rgba(16,24,40,.08), 0 6px 16px rgba(16,24,40,.06);
    --focus:0 0 0 3px rgba(32,86,214,.22);
  }
  * { box-sizing:border-box; }
  body { margin:0; font:15px/1.55 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
         background:var(--bg); color:var(--text); -webkit-font-smoothing:antialiased; }
  header { background:var(--card); border-bottom:1px solid var(--line); padding:22px 28px; }
  h1 { margin:0 0 3px; font-size:19px; letter-spacing:-.01em; }
  .gen { color:var(--muted); font-size:13px; }
  .wrap { max-width:1120px; margin:0 auto; padding:24px 28px 72px; }
  .stat .n, .sc .sv, .metric .mv { font-variant-numeric:tabular-nums; }
  .cards { display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr));
           gap:12px; margin-bottom:22px; }
  .stat { background:var(--card); border:1px solid var(--line); border-radius:var(--radius);
          padding:13px 15px; box-shadow:var(--shadow); }
  .stat .n { font-size:25px; font-weight:700; letter-spacing:-.02em; }
  .stat .l { color:var(--muted); font-size:12.5px; }
  .controls { display:flex; flex-wrap:wrap; gap:9px; align-items:center; margin-bottom:18px; }
  input[type=search], select { height:38px; padding:0 12px; border:1px solid var(--line-strong);
         border-radius:var(--radius-sm); font-size:14px; background:var(--card); color:var(--text);
         font-family:inherit; transition:border-color .15s, box-shadow .15s; }
  input[type=search] { flex:1; min-width:220px; }
  input[type=search]:focus, select:focus { outline:none; border-color:var(--link); box-shadow:var(--focus); }
  select { appearance:none; -webkit-appearance:none; padding-right:34px; cursor:pointer;
         background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='8' viewBox='0 0 12 8'%3E%3Cpath fill='none' stroke='%23697587' stroke-width='1.6' stroke-linecap='round' stroke-linejoin='round' d='M1 1.5 6 6.5 11 1.5'/%3E%3C/svg%3E");
         background-repeat:no-repeat; background-position:right 12px center; }
  label.chk { display:inline-flex; gap:7px; align-items:center; height:38px; background:var(--card);
         border:1px solid var(--line-strong); border-radius:var(--radius-sm); padding:0 12px; font-size:13px;
         cursor:pointer; color:var(--muted); transition:border-color .15s, color .15s; user-select:none; }
  label.chk:hover { border-color:var(--faint); color:var(--text); }
  label.chk input { accent-color:var(--link); width:15px; height:15px; }
  .row { background:var(--card); border:1px solid var(--line); border-radius:var(--radius);
         margin-bottom:10px; overflow:hidden; box-shadow:var(--shadow); transition:box-shadow .15s; }
  .row[open] { box-shadow:var(--shadow-hover); }
  .row > summary { list-style:none; cursor:pointer; padding:13px 15px;
         display:flex; gap:11px; align-items:flex-start; }
  .row > summary::-webkit-details-marker { display:none; }
  .row > summary:hover { background:#fafbfc; }
  .chev-btn { cursor:pointer; flex:none; width:22px; height:22px; color:var(--faint);
         font-size:10px; line-height:1; transition:transform .18s, color .15s; user-select:none;
         display:inline-flex; align-items:center; justify-content:center; margin-top:2px; }
  .row[open] .chev-btn { transform:rotate(90deg); color:var(--muted); }
  .row-main { flex:1; min-width:0; }
  .row-line1 { display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
  .url { font-weight:600; overflow-wrap:anywhere; user-select:text; cursor:text;
         flex:1; min-width:0; }
  .row-actions { display:flex; gap:6px; flex-wrap:wrap; margin-top:9px; }
  .row-actions button { cursor:pointer; border:1px solid var(--line-strong); background:var(--card);
         border-radius:7px; padding:5px 11px; font-size:12px; color:var(--link);
         white-space:nowrap; font-weight:600; font-family:inherit; transition:.12s; }
  .row-actions button:hover { background:#f3f6fd; border-color:var(--link); }
  .row-actions button:active { transform:translateY(1px); }
  .row-badges { margin-top:9px; }
  .tag { font-size:10.5px; font-weight:700; text-transform:uppercase; letter-spacing:.04em;
         padding:3px 7px; border-radius:5px; white-space:nowrap; flex:none; }
  .tag.col { background:#eef1fb; color:#2a3ea8; }
  .tag.bad { background:#fdecec; color:var(--crit); }
  .tag.ok { background:#e9f7ef; color:var(--ok); }
  .tag.meta-status { text-transform:none; letter-spacing:0; font-weight:600; background:#eef0f3;
         color:#5b6472; display:inline-block; max-width:240px; overflow:hidden;
         text-overflow:ellipsis; white-space:nowrap; vertical-align:bottom; }
  .badges { display:flex; gap:6px; flex-wrap:wrap; }
  .b { font-size:11.5px; padding:3px 9px; border-radius:20px; font-weight:600; white-space:nowrap; }
  .b.crit{background:#fdecec;color:var(--crit)} .b.serious{background:#fdeee2;color:var(--serious)}
  .b.moderate{background:#fbf3da;color:var(--moderate)} .b.minor{background:#eef0f3;color:var(--minor)}
  .b.links{background:#fdecec;color:var(--crit)} .b.zero{background:#e9f7ef;color:var(--ok)}
  .b.ok{background:#e9f7ef;color:var(--ok)}
  .detail { padding:4px 16px 18px; border-top:1px solid var(--line); background:#fcfcfd; }
  .detail h3 { font-size:11.5px; text-transform:uppercase; letter-spacing:.05em;
         color:var(--muted); margin:18px 0 9px; font-weight:700; }
  .meta-grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(230px,1fr)); gap:7px 22px; }
  .meta-grid div { font-size:13px; overflow-wrap:anywhere; }
  .meta-grid b { color:var(--muted); font-weight:600; }
  .v { border-left:3px solid var(--line); padding:7px 0 7px 13px; margin:9px 0; }
  .v.crit{border-color:var(--crit)} .v.serious{border-color:var(--serious)}
  .v.moderate{border-color:var(--moderate)} .v.minor{border-color:var(--minor)}
  .v .vh { font-weight:600; }
  .v .vt { color:var(--muted); font-size:12px; font-family:ui-monospace,SFMono-Regular,Menlo,monospace;
           overflow-wrap:anywhere; margin-top:3px; }
  .v a { color:var(--link); font-size:12px; }
  .broken { font-size:13px; padding:6px 0; border-bottom:1px solid var(--line); }
  .broken .s { font-weight:700; color:var(--crit); margin-right:8px; }
  .broken .u, .unc .u { overflow-wrap:anywhere; }
  .empty { color:var(--muted); padding:40px; text-align:center; }
  .nowrap { white-space:nowrap; }
  .unc { font-size:13px; padding:6px 0; border-bottom:1px solid var(--line); color:var(--muted); }
  .unc .s { font-weight:700; margin-right:8px; }
  .note { color:var(--muted); font-size:12px; margin:-2px 0 7px; }
  /* faner */
  .tabs { display:inline-flex; gap:4px; margin-bottom:20px; background:#eceef2;
          border-radius:10px; padding:4px; max-width:100%; flex-wrap:wrap; }
  .tab { padding:8px 15px; border:none; background:transparent; border-radius:7px; cursor:pointer;
         font-size:13.5px; font-weight:600; color:var(--muted); font-family:inherit; transition:.12s; }
  .tab:hover { color:var(--text); }
  .tab.active { background:var(--card); color:var(--text); box-shadow:var(--shadow); }
  .hidden { display:none; }
  /* sammenligning */
  .cmp { background:var(--card); border:1px solid var(--line); border-radius:var(--radius);
         margin-bottom:10px; overflow:hidden; box-shadow:var(--shadow); transition:box-shadow .15s; }
  .cmp[open] { box-shadow:var(--shadow-hover); }
  .cmp > summary { list-style:none; cursor:pointer; padding:14px 16px; }
  .cmp > summary::-webkit-details-marker { display:none; }
  .cmp > summary:hover { background:#fafbfc; }
  .cmp .top { display:flex; gap:10px; align-items:center; margin-bottom:11px; }
  .cmp .rowurl { font-weight:600; overflow-wrap:anywhere; flex:1; min-width:0; }
  .cmp .rowurl a { color:var(--link); text-decoration:none; }
  .cmp .rowurl a:hover { text-decoration:underline; }
  .cmp .chev { color:var(--faint); font-size:11px; transition:transform .18s; flex:none; }
  .cmp[open] .chev { transform:rotate(90deg); }
  .verdict { font-size:10.5px; font-weight:700; text-transform:uppercase; letter-spacing:.04em;
             padding:3px 8px; border-radius:5px; white-space:nowrap; flex:none; }
  .verdict.good { background:#e9f7ef; color:var(--ok); }
  .verdict.bad { background:#fdecec; color:var(--crit); }
  .verdict.same { background:#eef0f3; color:var(--minor); }
  .metrics { display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:9px; }
  .metric { border:1px solid var(--line); border-radius:var(--radius-sm); padding:9px 11px;
            font-size:13px; background:var(--card); }
  .metric .ml { color:var(--muted); font-size:12px; }
  .metric .mv { font-weight:700; font-size:15px; margin-top:3px; letter-spacing:-.01em; }
  .metric .mv .arrow { font-size:13px; margin-left:6px; }
  .arrow.good { color:var(--ok); } .arrow.bad { color:var(--crit); } .arrow.same { color:var(--faint); }
  .miss { color:var(--faint); font-style:italic; }
  .cmp-detail { padding:2px 16px 18px; border-top:1px solid var(--line); background:#fcfcfd;
         display:grid; grid-template-columns:repeat(auto-fit,minmax(300px,1fr)); gap:20px; }
  .cmp-detail .side h4 { font-size:12px; text-transform:uppercase; letter-spacing:.05em;
         color:var(--muted); margin:16px 0 8px; }
  .cmp-detail .sd-h { font-size:11.5px; text-transform:uppercase; letter-spacing:.05em;
         color:var(--muted); font-weight:700; margin:14px 0 6px; }
  .cmp-detail .sd-meta div { font-size:13px; margin-bottom:2px; overflow-wrap:anywhere; }
  .cmp-detail .sd-meta b { color:var(--muted); font-weight:600; }
  /* SEO / GEO */
  .seo-item { font-size:13px; padding:4px 0 4px 23px; position:relative; overflow-wrap:anywhere; }
  .seo-item::before { position:absolute; left:0; top:4px; font-weight:700; }
  .seo-item.fail::before { content:'✕'; color:var(--crit); }
  .seo-item.warn::before { content:'!'; color:var(--moderate); }
  .seo-item.ok::before { content:'✓'; color:var(--ok); }
  .seo-item code { font-family:ui-monospace,SFMono-Regular,Menlo,monospace; background:#eef0f3;
         padding:1px 5px; border-radius:4px; font-size:12px; }
  .tip { font-size:13px; padding:9px 11px; background:#eef4ff; border:1px solid #d9e4fb;
         border-radius:var(--radius-sm); margin:7px 0; color:#1c3a6b; }
  .pill { display:inline-block; font-size:11.5px; padding:3px 9px; border-radius:20px;
          margin:2px 5px 2px 0; font-weight:600; }
  .pill.good { background:#e9f7ef; color:var(--ok); }
  .pill.bad { background:#fdecec; color:var(--crit); }
  .pill.neutral { background:#eef0f3; color:var(--minor); }
  /* nettsted */
  .site { background:var(--card); border:1px solid var(--line); border-radius:var(--radius);
          margin-bottom:14px; padding:18px 20px; box-shadow:var(--shadow); }
  .site h2 { font-size:16px; margin:0 0 12px; overflow-wrap:anywhere; letter-spacing:-.01em; }
  .site h3 { font-size:11.5px; text-transform:uppercase; letter-spacing:.05em;
             color:var(--muted); margin:16px 0 8px; font-weight:700; }
  .bots { display:flex; flex-wrap:wrap; gap:6px; }
  .sm { font-size:13px; overflow-wrap:anywhere; padding:2px 0; }
  /* scoreboard (gammel vs ny) */
  .scoreboard-h { font-size:12px; text-transform:uppercase; letter-spacing:.05em;
                  color:var(--muted); margin:0 0 10px; font-weight:700; }
  .score { display:grid; grid-template-columns:repeat(auto-fit,minmax(155px,1fr));
           gap:12px; margin-bottom:22px; }
  .sc { background:var(--card); border:1px solid var(--line); border-radius:var(--radius);
        padding:13px 15px; box-shadow:var(--shadow); }
  .sc .sl { color:var(--muted); font-size:12px; }
  .sc .sv { font-size:18px; font-weight:700; margin-top:5px; white-space:nowrap; letter-spacing:-.01em; }
  .sc .sv .g { color:var(--faint); font-weight:600; }
  .sc .sv .arrow { font-size:13px; margin-left:5px; }
  .sub { color:var(--muted); font-size:12px; margin-top:4px; line-height:1.35; }
  /* skjermbilde */
  .shot { display:block; border:1px solid var(--line); border-radius:var(--radius-sm);
          overflow:hidden; overflow-y:auto; max-height:420px; background:#fff; margin-bottom:6px; }
  .shot img { width:100%; display:block; }
  .shot-link { display:inline-block; font-size:12px; color:var(--link); text-decoration:none;
               font-weight:600; margin-bottom:6px; }
  .shot-link:hover { text-decoration:underline; }
  /* oppfølging */
  .toolbar { display:flex; gap:8px; align-items:center; margin-bottom:14px; flex-wrap:wrap; }
  .toolbar .count { font-size:13px; color:var(--muted); margin-right:auto; }
  .toolbar button, .toolbar label.btn { cursor:pointer; border:1px solid var(--line-strong);
        background:var(--card); border-radius:8px; padding:7px 12px; font-size:13px;
        font-weight:600; color:var(--text); font-family:inherit; }
  .toolbar button:hover, .toolbar label.btn:hover { background:#f3f6fd; border-color:var(--link); color:var(--link); }
  .flag-btn { cursor:pointer; border:1px solid var(--line-strong); background:var(--card);
        border-radius:7px; padding:5px 11px; font-size:12px; font-weight:600; color:var(--moderate);
        font-family:inherit; white-space:nowrap; }
  .flag-btn:hover { background:#fbf3da; }
  .flag-btn.on { background:#fbf3da; border-color:var(--moderate); }
  .st-btn { cursor:pointer; border:1px solid var(--line-strong); background:var(--card);
        border-radius:7px; padding:5px 11px; font-size:12px; font-weight:600;
        font-family:inherit; white-space:nowrap; color:var(--muted); }
  .st-btn.flag:hover { background:#fbf3da; }
  .st-btn.flag.on { background:#fbf3da; border-color:var(--moderate); color:var(--moderate); }
  .st-btn.done:hover { background:#e9f7ef; }
  .st-btn.done.on { background:#e9f7ef; border-color:var(--ok); color:var(--ok); }
  .row.flagged { border-left:3px solid #d39a00; }
  .row.followup { border-left:3px solid #d39a00; }
  .row.done { border-left:3px solid var(--ok); }
  textarea.note { width:100%; min-height:52px; border:1px solid var(--line-strong); border-radius:8px;
        padding:8px 10px; font-family:inherit; font-size:13px; resize:vertical; background:var(--card);
        color:var(--text); }
  textarea.note:focus { outline:none; border-color:var(--link); box-shadow:var(--focus); }
</style>
</head>
<body>
<header>
  <h1>Validerings­rapport</h1>
  <div class="gen">Generert __GENERATED__</div>
</header>
<div class="wrap">
  <div class="toolbar">
    <span class="count" id="flag_count"></span>
    <button onclick="exportFollowup()">⬇ Eksporter oppfølging (JSON)</button>
    <label class="btn">Importer<input type="file" accept="application/json" style="display:none" onchange="importFollowup(this.files[0])"></label>
  </div>
  <div class="tabs">
    <button class="tab active" id="tab_cmp">Sammenlign (gammel vs ny)</button>
    <button class="tab" id="tab_list">Per side</button>
    <button class="tab" id="tab_site">Nettsted (robots / llms / AI)</button>
  </div>

  <!-- SAMMENLIGNING -->
  <div id="view_cmp">
    <div class="scoreboard-h">Samlet endring: gammel → ny</div>
    <div class="score" id="scoreboard"></div>
    <div class="controls">
      <input type="search" id="cq" placeholder="Søk i URL …">
      <select id="csort">
        <option value="row">Sorter: rekkefølge</option>
        <option value="improved">Sorter: mest forbedret</option>
        <option value="regressed">Sorter: mest forverret</option>
      </select>
      <label class="chk"><input type="checkbox" id="c_changed"> Bare endrede</label>
    </div>
    <div id="compare"></div>
  </div>

  <!-- PER SIDE -->
  <div id="view_list" class="hidden">
    <div class="controls">
      <input type="search" id="q" placeholder="Søk i URL …">
      <select id="col">
        <option value="">Alle kolonner</option>
        <option value="gammel">Gammel (url)</option>
        <option value="ny">Ny (ny-url)</option>
      </select>
      <select id="sort">
        <option value="row">Sorter: rekkefølge</option>
        <option value="a11y">Sorter: flest a11y-brudd</option>
        <option value="links">Sorter: flest brutte lenker</option>
      </select>
      <label class="chk"><input type="checkbox" id="f_a11y"> Bare a11y-feil</label>
      <label class="chk"><input type="checkbox" id="f_links"> Bare brutte lenker</label>
      <label class="chk"><input type="checkbox" id="f_load"> Bare lastefeil</label>
      <select id="f_status">
        <option value="">Status: alle</option>
        <option value="followup">Følg opp</option>
        <option value="done">Ferdig</option>
        <option value="none">Ikke vurdert</option>
      </select>
      <select id="f_seo">
        <option value="">SEO: alle</option>
        <option value="desc-missing">Mangler description</option>
        <option value="desc-long">Description for lang</option>
        <option value="desc-short">Description for kort</option>
        <option value="title-missing">Mangler title</option>
        <option value="og-image">Mangler og:image</option>
        <option value="canonical-missing">Mangler canonical</option>
        <option value="noindex">noindex</option>
        <option value="h1-none">Ingen h1</option>
        <option value="h1-many">Flere h1</option>
        <option value="heading-skips">Hopp i overskrifter</option>
      </select>
    </div>
    <div class="score" id="list_cards"></div>
    <div id="list"></div>
  </div>

  <!-- NETTSTED -->
  <div id="view_site" class="hidden">
    <div class="score" id="site_summary"></div>
    <div id="sites"></div>
  </div>
</div>
<script>
const DATA = __DATA__;
const esc = s => (s==null?"":String(s)).replace(/[&<>"]/g, c =>
  ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));

/* ---------- OPPFØLGING (status + notat, lagres i localStorage) ---------- */
const STORE_KEY = 'validering_followup:' + (DATA.source || '');
let FOLLOWUP = {};
try { FOLLOWUP = JSON.parse(localStorage.getItem(STORE_KEY) || '{}') || {}; } catch(e) { FOLLOWUP = {}; }
function saveFollowup(){ try { localStorage.setItem(STORE_KEY, JSON.stringify(FOLLOWUP)); } catch(e){} }
function fkey(p){ return p.row + '|' + p.column; }
function getMark(key){ return FOLLOWUP[key] || {status:'', note:''}; }
function setMark(key, patch){
  const m = Object.assign({status:'', note:''}, FOLLOWUP[key] || {}, patch);
  if (!m.status && !(m.note||'').trim()) delete FOLLOWUP[key]; else FOLLOWUP[key] = m;
  saveFollowup();
}
function applyRowStatus(el, st){
  if (!el) return;
  el.classList.remove('followup','done');
  if (st) el.classList.add(st);
}
function syncStatusUI(key){
  const st = getMark(key).status;
  document.querySelectorAll('.st-btn').forEach(btn=>{
    if (btn.dataset.stkey !== key) return;
    btn.classList.toggle('on', btn.dataset.st === st);
    applyRowStatus(btn.closest('.row'), st);
  });
}
function setStatus(ev, key, target){
  ev.stopPropagation(); ev.preventDefault();
  const cur = getMark(key).status;
  setMark(key, {status: cur === target ? '' : target});
  syncStatusUI(key); updateFlaggedUI();
}
function setNote(key, val){
  const cur = getMark(key);
  const patch = {note: val};
  if (val.trim() && !cur.status) patch.status = 'followup';  // notat auto-markerer
  setMark(key, patch);
  syncStatusUI(key); updateFlaggedUI();
}
function countStatus(s){ return Object.values(FOLLOWUP).filter(m=>m.status===s).length; }
function updateFlaggedUI(){
  const el = document.getElementById('flag_count');
  if (el) el.textContent = `🚩 ${countStatus('followup')} følges opp · ✓ ${countStatus('done')} ferdig`;
  if (window.__lastList) renderListCards(window.__lastList);
}
function exportFollowup(){
  const blob = new Blob([JSON.stringify(FOLLOWUP, null, 2)], {type:'application/json'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'oppfolging.json';
  document.body.appendChild(a); a.click(); a.remove();
  setTimeout(()=>URL.revokeObjectURL(a.href), 1000);
}
function importFollowup(file){
  if (!file) return;
  const r = new FileReader();
  r.onload = () => {
    try { FOLLOWUP = JSON.parse(r.result) || {}; saveFollowup();
          render(); renderCompare(); updateFlaggedUI(); }
    catch(e){ alert('Kunne ikke lese JSON-fila.'); }
  };
  r.readAsText(file);
}
function statusButtons(key){
  const st = getMark(key).status;
  return `<button class="st-btn flag ${st==='followup'?'on':''}" data-st="followup" data-stkey="${key}" onclick="setStatus(event,'${key}','followup')">⚑ Følg opp</button>
    <button class="st-btn done ${st==='done'?'on':''}" data-st="done" data-stkey="${key}" onclick="setStatus(event,'${key}','done')">✓ Ferdig</button>`;
}
function followupBlock(p){
  const key = fkey(p); const m = getMark(key);
  return `<h3>Oppfølging</h3>
    <div style="margin-bottom:8px">${statusButtons(key)}</div>
    <textarea class="note" oninput="setNote('${key}',this.value)"
      placeholder="Notat (lagres automatisk – markerer som «følg opp») …">${esc(m.note||'')}</textarea>`;
}
function extraBlock(p){
  const ex = p.extra || {};
  const keys = Object.keys(ex).filter(k => String(ex[k]).trim());
  if (!keys.length) return '';
  return `<h3>Info fra regneark</h3><div class="meta-grid">` +
    keys.map(k=>`<div><b>${esc(k)}:</b> ${esc(ex[k])}</div>`).join('') + `</div>`;
}
function statusTag(p){
  const ex = p.extra || {};
  const k = Object.keys(ex).find(k => k.toLowerCase().includes('status') && String(ex[k]).trim());
  return k ? `<span class="tag meta-status" title="${esc(ex[k])}">${esc(ex[k])}</span>` : '';
}

function a11yCount(p){ return p.a11y && p.a11y.violation_count || 0; }
function linkCount(p){ return p.links && p.links.broken ? p.links.broken.length : 0; }
function hasLoadErr(p){ return !!p.load_error || (p.status && p.status>=400); }
function jsonldCount(p){ return (p.meta && p.meta.jsonld || []).length; }

// par gammel/ny per rad, for å hente motpart-URL
const PAIR = {};
for (const p of DATA.pages){ (PAIR[p.row] = PAIR[p.row] || {})[p.column] = p; }
function counterpart(p){
  const o = PAIR[p.row] || {};
  return p.column==='ny' ? o.gammel : o.ny;
}

function fallbackCopy(t, done){
  const ta = document.createElement('textarea');
  ta.value = t; ta.style.position='fixed'; ta.style.opacity='0';
  document.body.appendChild(ta); ta.focus(); ta.select();
  try { document.execCommand('copy'); done(); } catch(e){}
  document.body.removeChild(ta);
}
function copyText(t, btn){
  const done = ()=>{ const o=btn.textContent; btn.textContent='Kopiert!';
                     setTimeout(()=>btn.textContent=o, 1200); };
  if (navigator.clipboard && navigator.clipboard.writeText)
    navigator.clipboard.writeText(t).then(done).catch(()=>fallbackCopy(t, done));
  else fallbackCopy(t, done);
}

function agg(pages){
  const a = {n:pages.length, loadErr:0, a11y:0, crit:0, serious:0, moderate:0, minor:0,
             incomplete:0, broken:0, uncertain:0, alt:0, seoFail:0, seoWarn:0,
             jsDep:0, md:0, jsonld:0, kbNoFocusPages:0, kbNoFocusEl:0, kbTrap:0};
  for (const p of pages){
    if (hasLoadErr(p)) a.loadErr++;
    a.a11y += a11yCount(p);
    const bi = (p.a11y && p.a11y.by_impact) || {};
    a.crit += bi.critical||0; a.serious += bi.serious||0;
    a.moderate += bi.moderate||0; a.minor += bi.minor||0;
    a.incomplete += (p.a11y && p.a11y.incomplete_count) || 0;
    a.broken += linkCount(p);
    a.uncertain += (p.links && p.links.uncertain ? p.links.uncertain.length : 0);
    a.alt += (p.meta && p.meta.images_missing_alt) || 0;
    for (const i of (p.seo||[])){ if(i.level==='fail') a.seoFail++; else if(i.level==='warn') a.seoWarn++; }
    const s = (p.geo && p.geo.signals) || {};
    if (s.js_dependent) a.jsDep++;
    if (s.markdown_available) a.md++;
    if (jsonldCount(p)) a.jsonld++;
    const k = p.keyboard;
    if (k){ if(k.no_focus_count){ a.kbNoFocusPages++; a.kbNoFocusEl += k.no_focus_count; } if(k.trap) a.kbTrap++; }
  }
  return a;
}

// betterLower=true: ned er bra (færre feil). false: opp er bra (mer dekning).
function arrow(oldN, newN, betterLower=true){
  if (oldN==null || newN==null) return '';
  const d = newN - oldN;
  if (d===0) return `<span class="arrow same">–</span>`;
  const improved = betterLower ? d<0 : d>0;
  const sym = d<0 ? '▼' : '▲';
  return `<span class="arrow ${improved?'good':'bad'}">${sym} ${Math.abs(d)}</span>`;
}

function scoreCard(label, g, n, betterLower=true){
  return `<div class="sc"><div class="sl">${label}</div>
    <div class="sv"><span class="g">${g}</span> → ${n} ${arrow(g,n,betterLower)}</div></div>`;
}

function byCol(c){ return DATA.pages.filter(p=>p.column===c); }

function renderScoreboard(){
  const G = agg(byCol('gammel')), N = agg(byCol('ny'));
  const cards = [
    scoreCard('A11y-brudd', G.a11y, N.a11y, true),
    scoreCard('— hvorav critical', G.crit, N.crit, true),
    scoreCard('— hvorav serious', G.serious, N.serious, true),
    scoreCard('Brutte lenker', G.broken, N.broken, true),
    scoreCard('Bilder uten alt', G.alt, N.alt, true),
    scoreCard('SEO-avvik', G.seoFail+G.seoWarn, N.seoFail+N.seoWarn, true),
    scoreCard('JS-avhengige sider', G.jsDep, N.jsDep, true),
    scoreCard('Sider m/ usynlig fokus', G.kbNoFocusPages, N.kbNoFocusPages, true),
    scoreCard('Sider m/ JSON-LD', G.jsonld, N.jsonld, false),
    scoreCard('Sider m/ markdown', G.md, N.md, false),
  ];
  document.getElementById('scoreboard').innerHTML = cards.join('');
}

function statCard(label, value, sub){
  return `<div class="sc"><div class="sl">${label}</div>
    <div class="sv">${value}</div>${sub?`<div class="sub">${sub}</div>`:''}</div>`;
}

function renderListCards(pages){
  const a = agg(pages);
  const sevSub = `${a.crit} critical · ${a.serious} serious · ${a.moderate} moderate · ${a.minor} minor`;
  const cards = [
    statCard('Sider (filtrert)', `${a.n} <span class="g" style="font-size:13px;color:var(--muted)">/ ${DATA.pages.length}</span>`),
    statCard('Lastefeil', a.loadErr),
    statCard('A11y-brudd', a.a11y, sevSub),
    statCard('Krever manuell sjekk', a.incomplete),
    statCard('Brutte lenker', a.broken, a.uncertain?`+ ${a.uncertain} uverifiserbare`:''),
    statCard('Bilder uten alt', a.alt),
    statCard('SEO-avvik', a.seoFail+a.seoWarn, `${a.seoFail} feil · ${a.seoWarn} advarsler`),
    statCard('JS-avhengige', a.jsDep),
    statCard('Usynlig fokus', a.kbNoFocusEl, a.kbTrap?`${a.kbTrap} mulige tab-feller`:'elementer ved tastaturfokus'),
    statCard('Følges opp', pages.filter(p=>getMark(fkey(p)).status==='followup').length),
    statCard('Ferdig', pages.filter(p=>getMark(fkey(p)).status==='done').length),
  ];
  document.getElementById('list_cards').innerHTML = cards.join('');
}

function renderSiteSummary(){
  const sites = Object.values(DATA.sites||{});
  let botsBlocked = 0, robotsMissing = 0, llms = 0;
  for (const s of sites){
    const r = s.robots||{};
    if (r.ai_bots) botsBlocked += Object.values(r.ai_bots).filter(v=>v===false).length;
    if (!r.exists) robotsMissing++;
    if (s.llms_txt && s.llms_txt.exists) llms++;
  }
  const cards = [
    statCard('Nettsteder', sites.length),
    statCard('Uten robots.txt', robotsMissing),
    statCard('AI-bot-blokkeringer', botsBlocked, 'på tvers av nettsteder'),
    statCard('Med llms.txt', `${llms} / ${sites.length}`),
  ];
  document.getElementById('site_summary').innerHTML = cards.join('');
}

function badges(p){
  const bi = (p.a11y && p.a11y.by_impact) || {};
  const out = [];
  for (const k of ['critical','serious','moderate','minor']){
    if (bi[k]) out.push(`<span class="b ${k.slice(0,4)==='crit'?'crit':k}">${bi[k]} ${k}</span>`);
  }
  const lc = linkCount(p);
  if (lc) out.push(`<span class="b links">${lc} brutte lenker</span>`);

  // Spesifikke SEO-badges: feil i rødt, utvalgte advarsler i gult, resten aggregert
  const SEO_LABEL = {
    'title-missing':'mangler title','title-long':'title for lang','title-short':'title for kort',
    'desc-missing':'mangler description','desc-long':'description for lang','desc-short':'description for kort',
    'lang-missing':'mangler lang','canonical-missing':'mangler canonical','viewport-missing':'mangler viewport',
    'h1-none':'ingen h1','h1-many':'flere h1','heading-skips':'hopp i overskrifter',
    'noindex':'noindex','nofollow':'nofollow','og-image':'mangler og:image','og-meta':'mangler OG',
    'twitter-missing':'mangler twitter:card'
  };
  const SEO_PROMOTE = new Set(['desc-missing','desc-long','desc-short','og-image',
                               'canonical-missing','h1-many','heading-skips']);
  let otherWarn = 0, shown = 0;
  for (const i of (p.seo||[])){
    if (i.level==='ok') continue;
    const label = SEO_LABEL[i.key] || 'SEO';
    if (i.level==='fail' && shown<5){ out.push(`<span class="b crit">${label}</span>`); shown++; }
    else if (i.level==='warn' && SEO_PROMOTE.has(i.key) && shown<5){ out.push(`<span class="b moderate">${label}</span>`); shown++; }
    else otherWarn++;
  }
  if (otherWarn) out.push(`<span class="b minor">+${otherWarn} SEO</span>`);

  const s = (p.geo && p.geo.signals) || {};
  if (s.js_dependent) out.push(`<span class="b moderate">krever JS</span>`);
  const k = p.keyboard;
  if (k && k.trap) out.push(`<span class="b crit">tab-felle</span>`);
  else if (k && k.no_focus_count) out.push(`<span class="b moderate">fokus usynlig (${k.no_focus_count})</span>`);
  if (p.load_error || (p.status && p.status>=400))
    out.push(`<span class="b crit">lastefeil</span>`);

  if (!out.length) out.push(`<span class="b ok">alt OK ✓</span>`);
  return `<div class="badges">${out.join('')}</div>`;
}

function seoBlock(p){
  const items = p.seo || [];
  const m = p.meta || {};
  let h = `<h3>SEO (${items.filter(i=>i.level!=='ok').length} ting å se på)</h3>`;
  if (!items.length) return h + '<div class="miss">Ingen data</div>';
  h += items.map(i=>`<div class="seo-item ${i.level}">${esc(i.msg)}</div>`).join('');
  // ekstra signaler
  const jl = (m.jsonld||[]);
  const hl = (m.hreflang||[]);
  h += `<div style="margin-top:8px">`;
  h += jl.length ? `<span class="pill good">JSON-LD: ${jl.map(esc).join(', ')}</span>`
                 : `<span class="pill bad">Ingen JSON-LD</span>`;
  if (hl.length) h += `<span class="pill neutral">hreflang: ${hl.length}</span>`;
  h += `</div>`;
  return h;
}

function geoBlock(p){
  const g = p.geo || {};
  const s = g.signals || {};
  let h = `<h3>GEO / AI-synlighet</h3>`;
  h += `<div>`;
  h += s.js_dependent ? `<span class="pill bad">Innhold krever JS</span>`
                      : `<span class="pill good">Server-rendret innhold</span>`;
  h += (s.markdown_available) ? `<span class="pill good">Markdown tilgjengelig</span>`
                             : `<span class="pill bad">Ingen markdown</span>`;
  if (s.word_count!=null) h += `<span class="pill neutral">${s.word_count} ord</span>`;
  if (s.ai_bots_blocked && s.ai_bots_blocked.length)
    h += `<span class="pill bad">AI-botter blokkert</span>`;
  h += `</div>`;
  // markdown-detaljer
  if (p.markdown){
    const md = p.markdown;
    const fmt = o => o ? (o.error ? 'feil: '+esc(o.error)
              : `${o.status} ${esc(o.content_type||'')}`) : '–';
    h += `<div class="note" style="margin-top:6px">
      Accept: text/markdown → ${fmt(md.accept_header)} &nbsp;|&nbsp; .md → ${fmt(md.dot_md)}</div>`;
  }
  const tips = g.tips || [];
  if (tips.length) h += tips.map(t=>`<div class="tip">💡 ${esc(t)}</div>`).join('');
  else h += `<div class="miss" style="margin-top:6px">Ingen åpenbare GEO-forbedringer 🎉</div>`;
  return h;
}

function a11yIncomplete(p){
  const inc = (p.a11y && p.a11y.incomplete) || [];
  if (!inc.length) return '';
  let h = `<h3>Krever manuell sjekk (${p.a11y.incomplete_count})</h3>`;
  h += `<div class="note">axe kunne ikke avgjøre disse automatisk – verdt en titt.</div>`;
  h += inc.map(i=>`<div class="seo-item warn">${esc(i.help)} (${i.nodes})
      <a href="${esc(i.helpUrl)}" target="_blank" rel="noopener" style="margin-left:6px;font-size:12px">info →</a></div>`).join('');
  return h;
}

function keyboardBlock(p){
  const k = p.keyboard;
  if (!k) return '';
  let h = `<h3>Tastatur / fokus</h3>`;
  h += `<div class="note">Tabber gjennom siden og sjekker synlig fokus-markering ved tastaturfokus (WCAG 2.4.7). «Usynlig fokus» = verken outline eller box-shadow.</div>`;
  h += `<div>`;
  h += `<span class="pill neutral">${k.tab_stops} tab-stopp</span>`;
  if (k.trap) h += `<span class="pill bad">Mulig tab-felle</span>`;
  if (k.skip_link) h += k.skip_link.present ? `<span class="pill good">Skip-lenke ✓</span>`
                                            : `<span class="pill neutral">Ingen skip-lenke</span>`;
  h += k.no_focus_count ? `<span class="pill bad">${k.no_focus_count} uten synlig fokus</span>`
                        : `<span class="pill good">Fokus synlig</span>`;
  h += `</div>`;
  if (k.no_focus && k.no_focus.length){
    h += `<div class="sd-h" style="margin-top:8px">Elementer uten synlig fokus</div>`;
    h += k.no_focus.map(e=>`<div class="seo-item warn">${esc(e.tag)} <code style="font-size:12px">${esc(e.sel)}</code>${e.name?' – «'+esc(e.name)+'»':''}</div>`).join('');
  }
  if (k.positive_tabindex && k.positive_tabindex.length){
    h += `<div class="sd-h" style="margin-top:8px">Positiv tabindex (forstyrrer rekkefølge)</div>`;
    h += k.positive_tabindex.map(e=>`<div class="seo-item warn">tabindex="${esc(e.tabindex)}" – <code style="font-size:12px">${esc(e.sel)}</code></div>`).join('');
  }
  if (k.unreachable_count){
    h += `<div class="sd-h" style="margin-top:8px">Interaktive elementer uten tastaturtilgang (${k.unreachable_count})</div>`;
    h += `<div class="note">role=button/link eller onclick uten tabindex – kan ikke nås med Tab.</div>`;
    h += (k.unreachable||[]).map(e=>`<div class="seo-item warn"><code style="font-size:12px">${esc(e.sel)}</code> (${esc(e.role)})${e.name?' – «'+esc(e.name)+'»':''}</div>`).join('');
  }
  if (k.aria_hidden_count){
    h += `<div class="sd-h" style="margin-top:8px">Fokuserbare elementer inni aria-hidden (${k.aria_hidden_count})</div>`;
    h += `<div class="note">I tab-rekkefølgen, men skjult for skjermlesere – forvirrende fokus.</div>`;
    h += (k.aria_hidden||[]).map(e=>`<div class="seo-item warn"><code style="font-size:12px">${esc(e.sel)}</code>${e.name?' – «'+esc(e.name)+'»':''}</div>`).join('');
  }
  if (k.offscreen_count){
    h += `<div class="sd-h" style="margin-top:8px">Fokuserbare utenfor skjerm (${k.offscreen_count})</div>`;
    h += `<div class="note">Kan være skjult innhold som fortsatt ligger i tab-rekkefølgen.</div>`;
  }
  return h;
}

function shotBlock(p){
  if (!p.shot) return '';
  return `<h3>Skjermbilde</h3>
    <div class="shot"><img src="${esc(p.shot)}" alt="Skjermbilde av ${esc(p.url)}" loading="lazy"></div>
    <a class="shot-link" href="${esc(p.shot)}" target="_blank" rel="noopener">Åpne i full størrelse ↗</a>`;
}

function detail(p){
  const m = p.meta || {};
  let h = '<div class="detail">';
  if (p.load_error) h += `<h3>Lastefeil</h3><div>${esc(p.load_error)}</div>`;
  h += followupBlock(p);
  h += extraBlock(p);
  h += shotBlock(p);
  h += `<h3>Meta</h3><div class="meta-grid">
    <div><b>HTTP-status:</b> ${p.status ?? '–'}</div>
    <div><b>Title (${m.title_length||0} tegn):</b> ${esc(m.title)||'<em>mangler</em>'}</div>
    <div><b>Description (${m.description_length||0} tegn):</b> ${esc(m.meta_description)||'<em>mangler</em>'}</div>
    <div><b>lang:</b> ${esc(m.lang)||'<em>mangler</em>'}</div>
    <div><b>h1 (antall):</b> ${m.h1_count ?? '–'}</div>
    <div><b>Bilder uten alt:</b> ${m.images_missing_alt ?? '–'} / ${m.images_total ?? '–'}</div>
    <div><b>Canonical:</b> ${esc(m.canonical)||'<em>mangler</em>'}</div>
    <div><b>robots-meta:</b> ${esc(m.robots_meta)||'<em>ingen</em>'}</div>
  </div>`;

  h += seoBlock(p);
  h += geoBlock(p);

  const vs = (p.a11y && p.a11y.violations) || [];
  h += `<h3>A11y-brudd (${vs.length})</h3>`;
  if (p.a11y && p.a11y.error) h += `<div>Feil under a11y-sjekk: ${esc(p.a11y.error)}</div>`;
  else if (!vs.length) h += `<div class="empty" style="padding:14px">Ingen brudd funnet 🎉</div>`;
  else h += vs.map(v=>`<div class="v ${v.impact.slice(0,4)==='crit'?'crit':v.impact}">
      <div class="vh">${esc(v.help)} <span style="color:var(--muted);font-weight:400">(${v.impact}, ${v.nodes} stk)</span></div>
      ${v.targets&&v.targets.length?`<div class="vt">${v.targets.map(esc).join(' · ')}</div>`:''}
      <a href="${esc(v.helpUrl)}" target="_blank" rel="noopener">Mer info →</a>
    </div>`).join('');
  h += a11yIncomplete(p);
  h += keyboardBlock(p);

  const br = (p.links && p.links.broken) || [];
  const unc = (p.links && p.links.uncertain) || [];
  const tot = p.links ? p.links.total : 0;
  h += `<h3>Brutte lenker: ${br.length} (av ${tot} sjekket)</h3>`;
  if (!br.length) h += `<div class="empty" style="padding:14px">Ingen brutte lenker 🎉</div>`;
  else h += br.map(b=>`<div class="broken">
      <span class="s">${b.status ?? 'FEIL'}</span>
      <span class="u">${esc(b.url)}</span>
      ${b.text?`<div style="color:var(--muted)">«${esc(b.text)}»</div>`:''}
    </div>`).join('');

  if (unc.length){
    h += `<h3>Kunne ikke verifiseres: ${unc.length}</h3>`;
    h += `<div class="note">Svarte 4xx/429/999 på automatisk forespørsel – ofte bot-blokkering eller krever innlogging. Sjekk gjerne manuelt.</div>`;
    h += unc.map(b=>`<div class="unc">
        <span class="s">${b.status ?? '?'}</span>
        <span class="u">${esc(b.url)}</span>
      </div>`).join('');
  }

  const ign = (p.links && p.links.ignored) || [];
  if (ign.length){
    h += `<h3>Ignorerte lenker: ${ign.length}</h3>`;
    h += `<div class="note">Manuelt ekskludert (forventet) – telles ikke som brutt.</div>`;
    h += ign.map(b=>`<div class="unc">
        <span class="s" style="color:var(--muted)">${b.status ?? '?'}</span>
        <span class="u">${esc(b.url)}</span>
      </div>`).join('');
  }

  h += '</div>';
  return h;
}

function render(){
  const q = document.getElementById('q').value.toLowerCase();
  const col = document.getElementById('col').value;
  const sort = document.getElementById('sort').value;
  const onlyA = document.getElementById('f_a11y').checked;
  const onlyL = document.getElementById('f_links').checked;
  const onlyE = document.getElementById('f_load').checked;
  const statusF = document.getElementById('f_status').value;
  const seoF = document.getElementById('f_seo').value;

  let pages = DATA.pages.filter(p=>{
    if (q && !p.url.toLowerCase().includes(q)) return false;
    if (col && p.column !== col) return false;
    if (onlyA && !a11yCount(p)) return false;
    if (onlyL && !linkCount(p)) return false;
    if (onlyE && !hasLoadErr(p)) return false;
    if (statusF){
      const st = getMark(fkey(p)).status || 'none';
      if (st !== statusF) return false;
    }
    if (seoF && !(p.seo||[]).some(i=>i.key===seoF && i.level!=='ok')) return false;
    return true;
  });
  if (sort==='a11y') pages = [...pages].sort((a,b)=>a11yCount(b)-a11yCount(a));
  else if (sort==='links') pages = [...pages].sort((a,b)=>linkCount(b)-linkCount(a));

  window.__lastList = pages;
  // Reaktive tellekort som speiler det aktive filteret
  renderListCards(pages);

  const list = document.getElementById('list');
  if (!pages.length){ list.innerHTML = '<div class="empty">Ingen sider matcher filteret.</div>'; return; }
  list.innerHTML = pages.map(p=>{
    const colTag = `<span class="tag col">${p.column==='ny'?'NY':'GAMMEL'}</span>`;
    const stTag = hasLoadErr(p) ? `<span class="tag bad nowrap">${p.status||'feil'}</span>`
                                : `<span class="tag ok nowrap">${p.status||'OK'}</span>`;
    const cp = counterpart(p);
    const cpBtn = cp ? `<button data-open="${esc(cp.url)}">${p.column==='ny'?'Åpne gammel ↗':'Åpne ny ↗'}</button>` : '';
    const key = fkey(p); const st = getMark(key).status;
    return `<details class="row${st?' '+st:''}">
      <summary>
        <span class="chev-btn" data-toggle aria-hidden="true">▶</span>
        <div class="row-main">
          <div class="row-line1">
            ${colTag}${stTag}${statusTag(p)}
            <span class="url">${esc(p.url)}</span>
          </div>
          <div class="row-actions">
            ${statusButtons(key)}
            <button data-open="${esc(p.url)}">Åpne ↗</button>
            <button data-copy="${esc(p.url)}">Kopier</button>
            ${cpBtn}
          </div>
          <div class="row-badges">${badges(p)}</div>
        </div>
      </summary>
      ${detail(p)}
    </details>`;
  }).join('');
}

/* ---------- SAMMENLIGNING (gammel vs ny) ---------- */
function pairRows(){
  const map = new Map();
  for (const p of DATA.pages){
    if (!map.has(p.row)) map.set(p.row, {row:p.row});
    map.get(p.row)[p.column] = p;
  }
  return [...map.values()].sort((a,b)=>a.row-b.row);
}
function score(p){ return p ? a11yCount(p)+linkCount(p) : null; }

function cell(label, oldN, newN){
  const o = oldN==null ? '<span class="miss">–</span>' : oldN;
  const n = newN==null ? '<span class="miss">–</span>' : newN;
  return `<div class="metric"><div class="ml">${label}</div>
    <div class="mv">${o} → ${n} ${arrow(oldN,newN)}</div></div>`;
}

function sideDetail(p){
  if (!p) return '<div class="miss">Ingen tilsvarende side</div>';
  const m = p.meta || {};
  const fk = fkey(p); const mark = getMark(fk);
  let h = `<div style="margin-bottom:8px">${statusButtons(fk)}</div>
    <textarea class="note" oninput="setNote('${fk}',this.value)" placeholder="Notat (markerer som «følg opp») …">${esc(mark.note||'')}</textarea>`;
  h += extraBlock(p);
  if (p.shot) h += `<div class="shot"><img src="${esc(p.shot)}" alt="Skjermbilde" loading="lazy"></div>
    <a class="shot-link" href="${esc(p.shot)}" target="_blank" rel="noopener">Åpne i full størrelse ↗</a>`;
  h += `<div class="sd-meta">
    <div><b>Status:</b> ${p.status ?? '–'}</div>
    <div><b>Title:</b> ${esc(m.title) || '<span class="miss">mangler</span>'}</div>
    <div><b>Description:</b> ${esc(m.meta_description) || '<span class="miss">mangler</span>'}</div>
    <div><b>lang:</b> ${esc(m.lang) || '<span class="miss">mangler</span>'} &nbsp; <b>h1:</b> ${m.h1_count ?? '–'} &nbsp; <b>Bilder uten alt:</b> ${m.images_missing_alt ?? '–'}/${m.images_total ?? '–'}</div>
  </div>`;

  const vs = (p.a11y && p.a11y.violations) || [];
  h += `<div class="sd-h">A11y-brudd (${vs.length})</div>`;
  if (p.a11y && p.a11y.error) h += `<div class="miss">Feil under sjekk: ${esc(p.a11y.error)}</div>`;
  else if (!vs.length) h += `<div class="miss">Ingen 🎉</div>`;
  else h += vs.map(v=>`<div class="v ${v.impact.slice(0,4)==='crit'?'crit':v.impact}">
      <div class="vh">${esc(v.help)} <span style="color:var(--muted);font-weight:400">(${v.impact}, ${v.nodes} stk)</span></div>
      ${v.targets&&v.targets.length?`<div class="vt">${v.targets.map(esc).join(' · ')}</div>`:''}
      <a href="${esc(v.helpUrl)}" target="_blank" rel="noopener">Mer info →</a>
    </div>`).join('');

  const br = (p.links && p.links.broken) || [];
  const unc = (p.links && p.links.uncertain) || [];
  h += `<div class="sd-h">Brutte lenker (${br.length})</div>`;
  if (!br.length) h += `<div class="miss">Ingen 🎉</div>`;
  else h += br.map(b=>`<div class="broken">
      <span class="s">${b.status ?? 'FEIL'}</span>
      <span class="u">${esc(b.url)}</span>
      ${b.text?`<div style="color:var(--muted)">«${esc(b.text)}»</div>`:''}
    </div>`).join('');
  if (unc.length){
    h += `<div class="sd-h">Kunne ikke verifiseres (${unc.length})</div>`;
    h += `<div class="note">Ofte bot-blokkering / krever innlogging.</div>`;
    h += unc.map(b=>`<div class="unc"><span class="s">${b.status ?? '?'}</span>
        <span class="u">${esc(b.url)}</span></div>`).join('');
  }

  // SEO-avvik (ikke ok)
  const seoIssues = (p.seo||[]).filter(i=>i.level!=='ok');
  h += `<div class="sd-h">SEO (${seoIssues.length} å se på)</div>`;
  if (!seoIssues.length) h += `<div class="miss">Alt OK 🎉</div>`;
  else h += seoIssues.map(i=>`<div class="seo-item ${i.level}">${esc(i.msg)}</div>`).join('');

  // GEO-tips
  const tips = (p.geo && p.geo.tips) || [];
  h += `<div class="sd-h">GEO / AI-synlighet</div>`;
  h += geoPills(p);
  if (tips.length) h += tips.map(t=>`<div class="tip">💡 ${esc(t)}</div>`).join('');
  else h += `<div class="miss">Ingen åpenbare forbedringer 🎉</div>`;

  return h;
}

function geoPills(p){
  const s = (p.geo && p.geo.signals) || {};
  let h = `<div>`;
  h += s.js_dependent ? `<span class="pill bad">Krever JS</span>`
                      : `<span class="pill good">Server-rendret</span>`;
  h += s.markdown_available ? `<span class="pill good">Markdown ✓</span>`
                           : `<span class="pill bad">Ingen markdown</span>`;
  const jl = (p.meta && p.meta.jsonld) || [];
  h += jl.length ? `<span class="pill good">JSON-LD ✓</span>`
                 : `<span class="pill bad">Ingen JSON-LD</span>`;
  h += `</div>`;
  return h;
}

function renderCompare(){
  const q = document.getElementById('cq').value.toLowerCase();
  const sort = document.getElementById('csort').value;
  const onlyChanged = document.getElementById('c_changed').checked;

  let rows = pairRows();
  rows.forEach(r=>{
    r.sOld = score(r.gammel); r.sNew = score(r.ny);
    r.diff = (r.sOld!=null && r.sNew!=null) ? r.sNew - r.sOld : 0;
  });
  // Vis bare rader som faktisk har både gammel og ny (et par å sammenligne)
  rows = rows.filter(r=>r.gammel && r.ny);
  if (q) rows = rows.filter(r=>{
    const u = (r.ny?.url || r.gammel?.url || '').toLowerCase();
    return u.includes(q);
  });
  if (onlyChanged) rows = rows.filter(r=>r.diff !== 0);
  if (sort==='improved') rows = [...rows].sort((a,b)=>a.diff-b.diff);
  else if (sort==='regressed') rows = [...rows].sort((a,b)=>b.diff-a.diff);

  const el = document.getElementById('compare');
  if (!rows.length){ el.innerHTML = '<div class="empty">Ingen par å sammenligne.</div>'; return; }

  el.innerHTML = rows.map(r=>{
    const g = r.gammel, n = r.ny;
    const url = (n && n.url) || (g && g.url) || '';
    let verdict = '<span class="verdict same">uendret</span>';
    if (r.diff < 0) verdict = '<span class="verdict good">forbedret</span>';
    else if (r.diff > 0) verdict = '<span class="verdict bad">forverret</span>';

    const gA = g?a11yCount(g):null, nA = n?a11yCount(n):null;
    const gL = g?linkCount(g):null, nL = n?linkCount(n):null;
    const gAlt = g&&g.meta?g.meta.images_missing_alt:null;
    const nAlt = n&&n.meta?n.meta.images_missing_alt:null;
    const combined = (n && n.combined_shot) || (g && g.combined_shot);

    return `<details class="cmp">
      <summary>
        <div class="top">
          ${verdict}
          <span class="rowurl"><a href="${esc(url)}" target="_blank" rel="noopener"
            onclick="event.stopPropagation()">${esc(url)}</a></span>
          <span class="chev">▶</span>
        </div>
        <div class="metrics">
          ${cell('A11y-brudd', gA, nA)}
          ${cell('Brutte lenker', gL, nL)}
          ${cell('Bilder uten alt', gAlt, nAlt)}
        </div>
      </summary>
      <div class="cmp-detail">
        ${combined?`<div style="grid-column:1/-1"><a class="shot-link" href="${esc(combined)}" target="_blank" rel="noopener">⬇ Last ned kombinert gammel/ny-bilde ↗</a></div>`:''}
        <div class="side"><h4>Gammel</h4>${sideDetail(g)}</div>
        <div class="side"><h4>Ny</h4>${sideDetail(n)}</div>
      </div>
    </details>`;
  }).join('');
}

/* ---------- NETTSTED (robots / llms / AI) ---------- */
function renderSites(){
  const sites = DATA.sites || {};
  const keys = Object.keys(sites);
  const el = document.getElementById('sites');
  if (!keys.length){ el.innerHTML = '<div class="empty">Ingen nettsted-data.</div>'; return; }

  el.innerHTML = keys.map(origin=>{
    const s = sites[origin];
    const r = s.robots || {};
    let h = `<div class="site"><h2>${esc(s.base || origin)}</h2>`;

    // robots.txt + AI-botter
    h += `<h3>robots.txt</h3>`;
    if (r.error) h += `<div class="miss">Feil: ${esc(r.error)}</div>`;
    else if (!r.exists) h += `<div class="miss">Ingen robots.txt (status ${r.status ?? '–'}) – alt er i praksis tillatt.</div>`;
    else {
      h += `<div class="note">Generelt (*): ${r.wildcard_allowed ? 'tillatt' : 'blokkert'}</div>`;
      h += `<h3>AI-botter</h3><div class="bots">`;
      for (const [bot,ok] of Object.entries(r.ai_bots||{})){
        const cls = ok===false ? 'bad' : (ok===true ? 'good' : 'neutral');
        const label = ok===false ? 'blokkert' : (ok===true ? 'tillatt' : '?');
        h += `<span class="pill ${cls}">${esc(bot)}: ${label}</span>`;
      }
      h += `</div>`;
      h += `<h3>Sitemaps (${(r.sitemaps||[]).length})</h3>`;
      h += (r.sitemaps&&r.sitemaps.length)
        ? r.sitemaps.map(u=>`<div class="sm"><a href="${esc(u)}" target="_blank" rel="noopener">${esc(u)}</a></div>`).join('')
        : `<div class="miss">Ingen sitemap oppgitt i robots.txt</div>`;
    }

    // llms.txt
    const llm = (key)=>{
      const o = s[key]; const path = key==='llms_txt'?'/llms.txt':'/llms-full.txt';
      if (!o || o.error) return `<span class="pill neutral">${path}: ukjent</span>`;
      return o.exists ? `<span class="pill good">${path}: finnes</span>`
                      : `<span class="pill bad">${path}: mangler (${o.status})</span>`;
    };
    h += `<h3>llms.txt</h3><div>${llm('llms_txt')} ${llm('llms_full_txt')}</div>`;
    h += `<div class="note">llms.txt er en framvoksende standard for å gi LLM-er et kuratert innholdskart. Valgfritt, men et pluss for AI-synlighet.</div>`;

    h += `</div>`;
    return h;
  }).join('');
}

/* ---------- faner ---------- */
function setTab(which){
  for (const t of ['cmp','list','site']){
    document.getElementById('tab_'+t).classList.toggle('active', which===t);
    document.getElementById('view_'+t).classList.toggle('hidden', which!==t);
  }
}

renderScoreboard();
render();
renderCompare();
renderSites();
renderSiteSummary();
updateFlaggedUI();

// Per side: bare chevron/tom plass åpner. URL/knapper gjør sin handling uten å toggle.
document.getElementById('list').addEventListener('click', (e)=>{
  const summary = e.target.closest('summary');
  if (!summary) return;
  const act = e.target.closest('[data-open],[data-copy],.url');
  if (!act) return;            // chevron eller tom plass -> la <details> toggle normalt
  e.preventDefault();          // hindre toggle for klikk på lenke/knapp
  if (act.hasAttribute('data-open'))
    window.open(act.getAttribute('data-open'), '_blank', 'noopener');
  else if (act.hasAttribute('data-copy'))
    copyText(act.getAttribute('data-copy'), act);
});

document.getElementById('tab_cmp').addEventListener('click', ()=>setTab('cmp'));
document.getElementById('tab_list').addEventListener('click', ()=>setTab('list'));
document.getElementById('tab_site').addEventListener('click', ()=>setTab('site'));
for (const id of ['q','col','sort','f_a11y','f_links','f_load','f_status','f_seo'])
  document.getElementById(id).addEventListener('input', render);
for (const id of ['cq','csort','c_changed'])
  document.getElementById(id).addEventListener('input', renderCompare);
</script>
</body>
</html>
"""


def build_html(data: dict) -> str:
    return HTML_TEMPLATE.replace(
        "__DATA__", json.dumps(data, ensure_ascii=False)
    ).replace("__GENERATED__", data.get("generated", ""))


async def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", type=Path, nargs="?", help="Sti til Excel-arket")
    parser.add_argument("--out", type=Path, default=Path("validering"),
                        help="Mappe for rapport (default: validering)")
    parser.add_argument("--only", choices=["url", "gammel", "ny"], default=None,
                        help="Sjekk bare én kolonne")
    parser.add_argument("--width", type=int, default=1440)
    parser.add_argument("--concurrency", type=int, default=3,
                        help="Antall sider samtidig (default 3)")
    parser.add_argument("--timeout", type=int, default=45000,
                        help="Maks lastetid per side i ms")
    parser.add_argument("--skip-links", action="store_true",
                        help="Hopp over lenkesjekk (raskere)")
    parser.add_argument("--internal-only", action="store_true",
                        help="Sjekk bare lenker på samme domene")
    parser.add_argument("--ignore-link", action="append", default=[], metavar="URL",
                        help="URL som IKKE skal telles som brutt (kan gjentas). "
                             "Avslutt med * for prefiks-treff, f.eks. https://x.no/api/*")
    parser.add_argument("--ignore-links-file", type=Path, default=None,
                        help="Fil med én ignorert-URL per linje")
    parser.add_argument("--fast", action="store_true",
                        help="Hopp over GEO/SSR/markdown/tastatur-sjekker (raskere)")
    parser.add_argument("--skip-keyboard", action="store_true",
                        help="Hopp over tastatur-/fokus-sjekken")
    parser.add_argument("--max-tabs", type=int, default=60,
                        help="Maks antall Tab-trykk per side i tastatur-sjekken")
    parser.add_argument("--screenshots", action="store_true",
                        help="Ta full-height skjermbilde av hver side og vis dem i rapporten")
    parser.add_argument("--rebuild", action="store_true",
                        help="Bygg report.html på nytt fra eksisterende report.json (ingen ny crawl)")
    parser.add_argument("--sitemap", type=str, default=None, metavar="URL",
                        help="Hent URL-er fra en sitemap.xml i stedet for Excel "
                             "(følger sitemap-index rekursivt). Hver URL blir én side.")
    parser.add_argument("--limit", type=int, default=None,
                        help="Maks antall sider å crawle (nyttig med --sitemap/--crawl)")
    parser.add_argument("--crawl", type=str, default=None, metavar="URL",
                        help="Fallback når sitemap mangler: traverser nettstedet fra denne "
                             "URL-en og følg interne lenker (samme origin).")
    args = parser.parse_args()

    # Rebuild-modus: bare regenerer HTML fra eksisterende JSON
    if args.rebuild:
        jpath = args.out / "report.json"
        if not jpath.exists():
            raise SystemExit(f"Fant ikke {jpath} – kjør en vanlig validering først.")
        data = json.loads(jpath.read_text(encoding="utf-8"))
        (args.out / "report.html").write_text(build_html(data), encoding="utf-8")
        print(f"Bygde report.html på nytt fra {jpath}.\nÅpne: {(args.out / 'report.html').resolve()}")
        return

    if args.sitemap:
        print(f"Henter URL-er fra sitemap: {args.sitemap}")
        pages = read_sitemap(args.sitemap, args.limit)
    elif args.crawl:
        print(f"Crawler nettstedet fra: {args.crawl}")
        pages = await crawl_site(args.crawl, args.limit)
    elif args.excel:
        pages = read_pages(args.excel, args.only)
        if args.limit and args.limit > 0:
            pages = pages[: args.limit]
    else:
        raise SystemExit(
            "Oppgi enten et Excel-ark, --sitemap <url> eller --crawl <url>, "
            "f.eks.: validate_pages.py --sitemap https://x.no/sitemap.xml --limit 20"
        )
    if not pages:
        raise SystemExit("Ingen gyldige URL-er funnet i arket.")

    # Samle ignorerte-lenke-mønstre fra flagg og evt. fil
    args.ignore_patterns = [p.strip() for p in (args.ignore_link or []) if p.strip()]
    if args.ignore_links_file:
        try:
            for line in args.ignore_links_file.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line and not line.startswith("#"):
                    args.ignore_patterns.append(line)
        except Exception as e:
            print(f"ADVARSEL: klarte ikke lese {args.ignore_links_file}: {e}")
    if args.ignore_patterns:
        print(f"Ignorerer {len(args.ignore_patterns)} lenke-mønster (telles ikke som brutt).")

    args.out.mkdir(parents=True, exist_ok=True)
    if args.screenshots:
        args.shots_dir = args.out / "shots"
        args.shots_dir.mkdir(exist_ok=True)
    axe_src = get_axe_source(args.out / "axe.min.js")

    link_cache: dict[str, int | None] = {}
    link_sem = asyncio.Semaphore(10)
    page_sem = asyncio.Semaphore(args.concurrency)
    results = [None] * len(pages)
    done = 0

    async with async_playwright() as p:
        browser = await p.chromium.launch()
        context = await browser.new_context(
            viewport={"width": args.width, "height": 1000},
            user_agent=UA,
            bypass_csp=True,
            ignore_https_errors=True,
        )

        # Nettsted-nivå (robots.txt / llms.txt) – én gang per unikt origin
        origins = {}
        for _, _, u, _ex in pages:
            pr = urlparse(u)
            origins.setdefault(pr.netloc, pr.scheme or "https")
        print(f"Sjekker {len(origins)} nettsted(er): robots.txt, llms.txt …")
        site_list = await asyncio.gather(
            *(check_site(context, o, s) for o, s in origins.items())
        )
        sites = {s["origin"]: s for s in site_list}

        async def work(i, row, column, url, extra):
            nonlocal done
            async with page_sem:
                origin = urlparse(url).netloc
                site_robots = (sites.get(origin) or {}).get("robots")
                entry = await analyze(
                    context, axe_src, row, column, url, args,
                    link_cache, link_sem, site_robots, extra
                )
                results[i] = entry
                done += 1
                v = entry["a11y"]["violation_count"] if entry.get("a11y") and "violation_count" in entry["a11y"] else 0
                b = len(entry["links"]["broken"]) if entry.get("links") else 0
                flag = "FEIL" if entry["load_error"] or (entry["status"] and entry["status"] >= 400) else "ok"
                print(f"[{done}/{len(pages)}] {flag}  a11y:{v} brutte:{b}  {url}")

        await asyncio.gather(
            *(work(i, r, c, u, ex) for i, (r, c, u, ex) in enumerate(pages))
        )
        await browser.close()

    # Sy sammen gammel + ny til ett bilde per rad (hvis skjermbilder er tatt)
    if args.screenshots:
        try:
            import PIL  # noqa: F401 – sjekk at Pillow finnes
            by_row = {}
            for e in results:
                if e:
                    by_row.setdefault(e["row"], {})[e["column"]] = e
            made = 0
            for row, o in by_row.items():
                g, n = o.get("gammel"), o.get("ny")
                if g and n and g.get("shot") and n.get("shot"):
                    out_p = args.shots_dir / f"{row}_sammenligning.jpg"
                    try:
                        stitch_pair(args.out / g["shot"], args.out / n["shot"], out_p)
                        rel = f"shots/{out_p.name}"
                        g["combined_shot"] = rel
                        n["combined_shot"] = rel
                        made += 1
                    except Exception as ex:
                        print(f"  kunne ikke sy sammen rad {row}: {ex}")
            print(f"Laget {made} kombinerte gammel/ny-bilder.")
        except ImportError:
            print("ADVARSEL: Pillow mangler – hopper over kombinerte bilder "
                  "(installer med: pip install pillow).")

    data = {
        "generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source": str(args.excel),
        "pages": results,
        "sites": sites,
    }
    (args.out / "report.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (args.out / "report.html").write_text(build_html(data), encoding="utf-8")

    print(f"\nFerdig. Åpne: {(args.out / 'report.html').resolve()}")


if __name__ == "__main__":
    asyncio.run(main())
